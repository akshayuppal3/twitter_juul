#############################
##Class for scraping ########
##following data twitter#####
##########version 2 #########
#############################

from authentication import Authenticate
from tweepy import OAuthHandler
import pandas as pd
import util
import argparse
import logging
import tweepy
import os
import openpyxl
from openpyxl import load_workbook


logging.basicConfig(level="INFO", format= util.format, filename= os.path.join(util.logdir,"followingData.log"))


## just passing the file and it would extract the following data

class twitter_following():

    def __init__(self):
        self.api = self.authorization()

    ## currently returning one API
    ## @TODO need to change for multiple API
    def authorization(self):
        ob = Authenticate()
        consumer_key = ob.getConsumerKey()
        consumer_secret = ob.getConsumerSecret()
        access_token = ob.getAccessToken()
        access_secret = ob.getAccessSecret()
        author = OAuthHandler(consumer_key, consumer_secret)
        author.set_access_token(access_token, access_secret)
        # change to accomodate rate limit errors
        api = tweepy.API(author, wait_on_rate_limit=True, wait_on_rate_limit_notify=True)
        return (api)

    # @param df, filename , testMode(bool)
    # @return None
    # writes to an excel file
    def getFriendsData(self,df,filename,testMode =False):
        path = os.getcwd()
        filepath = os.path.join(path,filename)
        users = util.getUsers(df,type= 'id')
        try:
            if users is not None:
                for user in users:
                    friendList = self.api.friends_ids(user,
                                                      count=util.friendLimit)  # returns list of friends
                    df = pd.DataFrame({'userID':user,
                                       'following':[friendList]}, index=[0])
                    writer = pd.ExcelWriter(filepath, engine='openpyxl')
                    if os.path.isfile(filepath):
                        writer.book = load_workbook(filepath)
                        writer.sheets = dict((ws.title, ws) for ws in writer.book.worksheets)
                        max_row = writer.book.active.max_row
                        sheetname = writer.book.active.title
                        df.to_excel(writer, sheet_name=sheetname, startrow=max_row, index=False, header=False)
                    else:
                        df.to_excel(writer, index=False)
                    try:
                        writer.save()
                    except OSError:
                        logging.error("File is open: or permission denied")

        except tweepy.TweepError as e:          # except for handling tweepy api call
            print("[Error] " + e.reason)

if __name__ == '__main__':
    ob = twitter_following()
    parser = argparse.ArgumentParser(description='Extracting data from twintAPI')
    parser.add_argument('-i', '--inputFile', help='Specify the input file path for extracting friends', required=True)
    parser.add_argument('-o',  '--outputFile', help='Specify the output file name with following data',default='followingList')
    args = vars(parser.parse_args())
    if (args['inputFile']):
        logging.info('[NEW] ---------------------------------------------')
        logging.info('new extraction process started')
        filename_input = args['inputFile']
        filename_output = args['outputFile']
        os.path.join(util.inputdir,filename_output)
        df = util.readCSV(filename_input)
        ob.getFriendsData(df,filename_output)
        logging.info("File creation completed")





