####################
##Class containing##
# helper functions###
####################
from time import sleep
import tweepy
import os
import argparse
import pandas as pd
import pandas.io.common
from pathlib import Path
import json
from openpyxl import load_workbook
from openpyxl.utils.exceptions import IllegalCharacterError
import posixpath
import numpy as np
import nltk
import ast
from setup import setup_env
from tqdm import tqdm
import git
import networkx as nx
from keras.models import model_from_json
from keras.preprocessing.sequence import pad_sequences
import pickle
tqdm.pandas()

setup_env()  # download necessary nltk packages
stopwords = nltk.corpus.stopwords.words('english')

## loading the config file
dir_name = os.getcwd()
path1 = str(Path(os.getcwd()).parent.parent.parent)
filepath = posixpath.join(path1, 'config.json')
with open(filepath) as f:
	data = json.load(f)
logdir = os.path.join(path1, data['logdir'])
twintDir = os.path.join(path1, data['twintdir'])
inputdir = os.path.join(path1, data['inputdir'])
modeldir = os.path.join(path1, data['modeldir'])
embeddir = os.path.join(path1, inputdir, 'embeddings')

## logging informaation
format = "%(asctime)s %(levelname)-8s %(message)s"
dateFormat = "%Y-%m-%d"
testLimit = 5
twintLimit = 1
userTimelineLimit = 200  # limit for the no of tweets extracted from user timeline
friendLimit = 100
startDate = '2018-05-01'
endDate = '2018-05-02'


def is_int(ele):
	try:
		int(ele)
		return True
	except ValueError:
		return False


def is_float(ele):
	try:
		float(ele)
		return True
	except ValueError:
		return False


def is_number(ele):
	try:
		if len(ele):
			return False
	except:
		if (is_int(ele) or is_float(ele)):
			return True
		else:
			return False  # can't be determined

# get the git repo
def get_git_root(path):
	git_repo = git.Repo(path, search_parent_directories=True)
	git_root = git_repo.git.rev_parse("--show-toplevel")
	return git_root

# @param dataframe and output filename
def output_to_csv(df, filename):
	if (df is not None and not df.empty):
		df.to_csv(filename, sep=",", line_terminator='\n', index=None)
	else:
		print("datframe is empty")


# conversion of str to bool
def str2bool(v):
	if v.lower() in ('yes', 'true', 't', 'y', '1'):
		return True
	elif v.lower() in ('no', 'false', 'f', 'n', '0'):
		return False
	else:
		raise argparse.ArgumentTypeError('Boolean value expected.')


# @Deprecated
# handle the rate limit (wait for 15min (API constarints))
def limit_handler(cursor):
	while True:
		try:
			yield cursor.next()
		except tweepy.RateLimitError:
			print("sleeping -rate limit exceeded")
			sleep(15)


# @param df
# @return list of unique users
def getUsers(df, type):
	if (df is not None and not df.empty):
		if (type == 'ID'):
			if ('userID' in df):
				unique_users = list(df['userID'].unique())
				return unique_users
			else:
				return None
		elif (type == 'name'):
			if ('userName' in df):
				unique_names = list(df['userName'].unique())
				return unique_names
			else:
				return None
	else:
		return None


# read CSV to generate dataframe
# @return df
def readCSV(path):
	try:
		df = pd.read_csv(path, lineterminator='\n', index_col=None)
		if "userName\r" in df:  # windows problem
			df["userName\r"] = df["userName\r"].str.replace(r'\r', '')
			df.rename(columns={'userName\r': "userName"}, inplace=True)
		return df
	except FileNotFoundError:
		print("[ERROR] file not found")
	except pd.io.common.EmptyDataError:
		print("[ERROR] empty file")


# read xlsx file as csv and return dataframe
# @return df @param filepath
def read_excel(path):
	try:
		df = pd.read_excel(path, 'Sheet1', index_col=None)
		return df
	except FileNotFoundError:
		print("[ERROR] file not found")
	except pd.io.common.EmptyDataError:
		print("[ERROR] empty file")


# Convert df to excel
# appends to the excel file path specified(or create a nex file with that name)
def df_write_excel(df, filepath):
	# df = df.applymap(lambda x: x.encode('unicode_escape').
	#                  decode('utf-8') if isinstance(x, str) else x)             # prevent Illegal character errror
	try:
		writer = pd.ExcelWriter(filepath, engine='openpyxl')
		if os.path.isfile(filepath):
			writer.book = load_workbook(filepath)
			writer.sheets = dict((ws.title, ws) for ws in writer.book.worksheets)
			max_row = writer.book.active.max_row
			sheetname = writer.book.active.title
			if len(df > 1):
				for index, row in df.iterrows():
					try:
						row.to_excel(writer, sheet_name=sheetname, startrow=max_row, index=True, header=False)
					except IllegalCharacterError:
						print("Illegal character error")
						continue
			else:
				df.to_excel(writer, sheet_name=sheetname, startrow=max_row, index=False, header=False)
		else:
			df.to_excel(writer, index=False)  # in case the file does not exists
		try:
			writer.save()
		except OSError:
			print("File is open: or permission denied")
	except IllegalCharacterError:
		print("Illegal character error")


# appends to existing csv file
def df_write_csv(df, filepath):
	try:
		if os.path.isfile(filepath):
			with open(filepath, 'a') as f:
				df.to_csv(f, header=False)
		else:
			df.to_csv(filepath)  # in case the file does not exists
	except IllegalCharacterError:
		print("Illegal character error")


## @param df  , @return length of hashtags
def hashtag_count(df):
	if df is not np.nan:
		hashtags = ast.literal_eval(df)
		if (hashtags is not None):
			return (len(hashtags))
		else:
			return (0)
	else:
		return (0)

## return tokenize sentences
# @param df, columns
# @returns sentences
def get_sentences(df, column):
	sentences = list(df[column].progress_apply(get_tokenize_words))
	return (sentences)


# function to return tokenize words
## @param text @return tokenize words
def get_tokenize_words(text):
	tkz = nltk.tokenize
	words = tkz.word_tokenize(text)
	words = [ele for ele in words if ((ele not in stopwords) and len(ele) > 2 and (ele.isalpha()))]
	words = [word.lower() for word in words]
	return words


# function to get the nearest postion for element in the list
def nearest(items, pivot):
	return min(items, key=lambda x: abs(x - pivot))


def getHashtags(tweetObj, extended=False):
	if extended == True:
		if 'retweeted_status' in tweetObj._json.keys():
			if len(tweetObj.retweeted_status.entities['hashtags']) != 0:
				hashtags = [i['text'] for i in tweetObj.retweeted_status.entities['hashtags']]
			else:
				hashtags = None
		else:
			hashtags = getHashtags(tweetObj, extended=False)
	else:
		if len(tweetObj.entities['hashtags']) != 0:
			hashtags = [j['text'] for j in tweetObj.entities['hashtags']]
		else:
			hashtags = "None"
	return hashtags

# @params passing tweet, friendOpt and userinfo(in case of following)
# returns data frame of tweet and user info
def getTweetObject(tweetObj, parentID=None):
	if parentID is not None:
		data = pd.DataFrame.from_records(
			[{
				'userID': tweetObj.id,
				'parentID': parentID,
				'userName': tweetObj.name,
				'userDescription': tweetObj.description,
				'userCreatedAt': tweetObj.created_at,
				'userLocation': tweetObj.location,
				'favourites_count': tweetObj.favourites_count,
				'friendsCount': tweetObj.friends_count,
				'userFollowersCount': tweetObj.followers_count,
				'listedCount': tweetObj.listed_count,
				'lang': tweetObj.lang,
				'url': tweetObj.url,
				'imageurl': tweetObj.profile_image_url,
				'userVerified': tweetObj.verified,
				'isProtected': tweetObj.protected,
				'notifications': tweetObj.notifications,
				'statusesCount': tweetObj.statuses_count,
				'geoEnabled': tweetObj.geo_enabled,
				'contributorEnabled': tweetObj.contributors_enabled,
				# 'status': tweetObj.status,
				'withheldinCountries': tweetObj.withheld_in_countries if 'withheld_in_countries' in tweetObj._json.keys() else None,
				'withheldScope': tweetObj.withheld_scope if 'withheld_scope' in tweetObj._json.keys() else None,
			}], index=[0])
	else:
		if 'retweeted_status' in tweetObj._json.keys():
			retweetText = tweetObj.retweeted_status.full_text.replace("\n", " ")
			text = tweetObj.full_text.replace("\n", " ")
			hashtags = getHashtags(tweetObj, extended=True)  # for retweeted status
			retweeted = True
		else:
			text = tweetObj.full_text.replace("\n", " ")
			retweetText = None
			hashtags = getHashtags(tweetObj)
			retweeted = False
		data = pd.DataFrame.from_records(
			[{
				'tweetId': tweetObj.id_str,
				'tweetText': text,
				'retweetText': retweetText,
				'retweetCount': tweetObj.retweet_count,
				'retweeted': retweeted,
				'tweetCreatedAt': tweetObj.created_at,
				'userName': tweetObj.user.name,
				'userID': tweetObj.user.id,
				'userCreatedAt': tweetObj.user.created_at,
				'userLocation': tweetObj.user.location,
				'userDescription': tweetObj.user.description.replace("\n", " "),
				'friendsCount': tweetObj.user.friends_count,
				'followersCount': tweetObj.user.followers_count,
				'listedCount': tweetObj.user.listed_count,
				'lang': tweetObj.lang,
				'hashtags': hashtags,
				'url': tweetObj.user.url,
				'imageurl': tweetObj.user.profile_image_url,
			}], index=[0])
	return (data)


### building the network
## @param : following network : columns <userID, following>
def get_graph(df : pd.DataFrame()) -> nx.DiGraph():
	G = nx.DiGraph()
	for user in tqdm(list(df.userID)):
		following_A = set(ast.literal_eval((df.loc[df.userID == user].head(1)["following"].values)[0])) ## get all of the following of user
		user_set = set([node for node in list(df.userID) if node != user])
		users_list = user_set.intersection(following_A)                     # intersect follwoing with the remaining users..
		for user_followed_by in list(users_list):
			G.add_edge(user, user_followed_by)
	return (G)

## dump the neural model and weights
def dump_model(model,path):
	# serialize model to JSON
	model_json = model.to_json()
	with open(path, "w") as json_file:
		json_file.write(model_json)
	# serialize weights to HDF5
	model.save_weights("model.h5")

## load the bilstm mode and weights
def load_model(path):
	# load json and create model
	json_file = open(path, 'r')
	loaded_model_json = json_file.read()
	json_file.close()
	loaded_model = model_from_json(loaded_model_json)
	return loaded_model

## file can be .pkl or .csv
def read_file(path):
	if path.endswith('.pkl'):
		df_input = pickle.load(open(path, "rb"))
	elif path.endswith('.csv'):
		df_input = pd.read_csv(path, lineterminator='\n')
	return df_input

def get_encoded_data(data_,tokenizer,max_len):
	tokenizer.fit_on_text(data_)
	encoded_docs = tokenizer.texts_to_sequences(data_)
	data = pad_sequences(encoded_docs,maxlen=max_len,padding='post')
	return data

def pickle_file(data,path):
	with open(path,"wb") as f:
		pickle.dump(data,f)