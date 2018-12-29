####################
##Class containing##
# helper functions###
####################
from time import sleep
import os
import argparse
import numpy as np
import pandas as pd
import pandas.io.common
from pathlib import Path
import json
import ast
from openpyxl import load_workbook
from openpyxl.utils.exceptions import IllegalCharacterError


dir_name = os.getcwd()
path1 = Path(os.getcwd()).parent.parent
filepath = os.path.join(path1, 'config.json')
with open(filepath) as f:
    data = json.load(f)
logdir = os.path.join(path1,data['logdir'])
inputdir = os.path.join(path1,data['inputdir'])
format = "%(asctime)s %(levelname)-8s %(message)s"


def hashtag_count(df):
    if df is not np.nan:
        hashtags = ast.literal_eval(df)
        if (hashtags is not None):
            return (len(hashtags))
        else:
            return (0)
    else:
        return (0)

# @param dataframe and output filename
def output_to_csv(df, filename):
    if (df is not None and not df.empty):
        df.to_csv(filename, sep=",", line_terminator='\n', index=None)
    else:
        print("datframe is empty")


# conversion of str to bool
def str2bool(v):
    if v.lower() in ('yes', 'true', 't', 'y', '1'):
        return True
    elif v.lower() in ('no', 'false', 'f', 'n', '0'):
        return False
    else:
        raise argparse.ArgumentTypeError('Boolean value expected.')

# Convert df to excel
# appends to the excel file path specified(or create a nex file with that name)
def df_write_excel(df,filepath):
    # df = df.applymap(lambda x: x.encode('unicode_escape').
    #                  decode('utf-8') if isinstance(x, str) else x)             # prevent Illegal character errror
    try:
        writer = pd.ExcelWriter(filepath, engine='openpyxl')
        if os.path.isfile(filepath):
            writer.book = load_workbook(filepath)
            writer.sheets = dict((ws.title, ws) for ws in writer.book.worksheets)
            max_row = writer.book.active.max_row
            sheetname = writer.book.active.title
            if len(df > 1):
                for index,row in df.iterrows():
                    try:
                        row.to_excel(writer, sheet_name=sheetname, startrow=max_row, index=False, header=False)
                    except IllegalCharacterError:
                        print(row)
                        print("Illegal character error")
                        continue
            else:
                df.to_excel(writer, sheet_name=sheetname, startrow=max_row, index=False, header=False)
        else:
            df.to_excel(writer, index=False)       #in case the file does not exists
        try:
            writer.save()
        except OSError:
            print("File is open: or permission denied")
    except IllegalCharacterError:
        print("Illegal character error")


